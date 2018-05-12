import os
from openpyxl import load_workbook
from pathlib import Path
import json
import pprint
import copy
import datetime
from dateutil.parser import parse
from django.conf import settings
from .dbConnect import dbConnect
from collections import namedtuple


class Timesheet:
    def __init__(self, workdir=os.getcwd()):
        self.__workdir = workdir
        self.__printFalg = False

    def getcurrdir(self):
        return self.__workdir

    def setcurrdir(dirname):
        self.__workdir = dirname

    def getWorkBookObj(self, filename):
        wb = load_workbook(filename)
        return wb

    def getSheetObj(self, wb, sheeetName):
        # return wb.get_sheet_by_name(sheeetName)
        return wb[sheeetName]

        # Get a sheet by name            

    def getSheetNames(self, wb):
        print(wb.sheetnames)
        #         sheet = wb.get_sheet_by_name('1,2,3')
        #         return wb.get_sheet_names()
        return wb.sheetnames

    def printData(self, sheet):
        # sheet.max_row
        # sheet.max_column
        for rwNum in range(2, sheet.max_row + 1):
            for colNum in range(1, sheet.max_column):
                if (sheet.cell(row=rwNum, column=colNum).value is not None):
                    #             print ( type(sheet.cell(row=rwNum, column=colNum).value) )
                    print(sheet.cell(row=rwNum, column=colNum).value, end=' ', sep=' ')
            print("\n")

    # load configuration data from the config file
    def loadconfigdata(self):
        cnfgfile = os.path.join(settings.CUSTOM_DIR, "cellmapping")

        with open(cnfgfile) as f:
            data = ''.join(line.rstrip() for line in f)
            d = json.loads(data)
        return d

    def parseDay(self, dayAbbr):
        returnStr = ""
        if "Mo" in dayAbbr:
            returnStr = "Mon"
        elif "Tu" in dayAbbr:
            returnStr = "Tue"
        elif "We" in dayAbbr:
            returnStr = "Wed"
        elif "Th" in dayAbbr:
            returnStr = "Thu"
        elif "Fr" in dayAbbr:
            returnStr = "Fri"
        elif "Sa" in dayAbbr:
            returnStr = "Sat"
        elif "Su" in dayAbbr:
            returnStr = "Sun"
        else:
            print("Str Not Found", dayAbbr)
            returnStr = dayAbbr
        self.__printFalg and print("returnStr", returnStr)
        return returnStr

    # get configuration mapping that defines the cells to be extracted, defined in the configuration file
    def getcellMapping(self, cnfgjson):
        cellcolrange = cnfgjson["attendance_data_range"]["colrange"]  # # array of range of cell mapping
        cellrwrangest = int(cnfgjson["attendance_data_range"]["rowrange"]["start"])  # # start of row for data capture
        cellrwrangeend = cnfgjson["attendance_data_range"]["rowrange"]["end"]  # # end of row for data capture   
        namelist = cnfgjson["name"]  ## list of employee names in the sheet
        currmonth = cnfgjson["date"]  ## get date specified in particular cell
        daysAbsent = cnfgjson["absence"]
        daysWorked = cnfgjson["daysWorked"]
        return cellcolrange, cellrwrangest, cellrwrangeend, namelist, currmonth, daysAbsent, daysWorked

    # get the data from the collection stored in the database and return the set,
    # this set will be used to check if the current data exists in the set or not
    def getDBShopConfigData(self):
        dbObj = dbConnect()
        db = dbObj.getDbConn("youandi")
        coll = dbObj.getCollection(db, "shopConfiguration")
        output = coll.find({}, {"_id": 0})
        print("getDBShopConfigData")
        shopConfigDates = {}
        # Shop Configuration information
        for doc in output:
            for key, val in doc.items():
                shopConfigDates[key] = val
                self.__printFalg and print(key, val)

        self.__printFalg and print(shopConfigDates)
        # Employee Information
        coll = dbObj.getCollection(db, "empDetails")
        output = coll.find({}, {"_id": 0})
        print("empDetails")
        empDetails = {}
        for doc in output:
            for key, val in doc.items():
                empDetails[key] = val
                self.__printFalg and print(key, val)

        self.__printFalg and print(shopConfigDates)
        self.__printFalg and print(empDetails)

        return shopConfigDates, empDetails

    # Massage data
    def massageData(self, inputString):
        massagedString = inputString.replace('+', '')
        return massagedString

    # loop through all the workbooks in the excel and read the contents fetched within parameters of the configuration file
    def getSheetData(self, filename):
        try:
            pp = pprint.PrettyPrinter(indent=4)
            wbObj = self.getWorkBookObj(filename)
            worksheets = self.getSheetNames(wbObj)

            cnfgjson = self.loadconfigdata()
            cellcolrange, cellrwrangest, cellrwrangeend, namelist, currmonth, daysAbsent, daysWorked = self.getcellMapping(
                cnfgjson)

            self.__printFalg and print("Starting1")
            self.__printFalg and print(cellcolrange, cellrwrangest, cellrwrangeend, namelist, currmonth)
            self.__printFalg and print("Ending1")

            extractedData = {}
            # Uncomment it after testing
            for sheet in range (2, len(worksheets) - 1, 1):
            # for sheet in range(2, 4, 1):
                self.__printFalg and print("Starting2")
                sheetName= None
                sheetName = self.getSheetObj(wbObj, worksheets[sheet])
                self.__printFalg and print("Ending2")
                print(sheetName , worksheets[sheet] ,cellcolrange[0][0], cellcolrange[0][1],  cellcolrange[1][0], cellcolrange[1][1],  cellcolrange[2][0], cellcolrange[2][1], "maxrw:", sheetName.max_row )
                cellrwrangeend = sheetName.max_row
                # extractedData[worksheets[sheet]]=[]
                puchtime = list()
                #             get cell value for month, split it on ~ take the first component of array and then split the date and then get just the month and year portion
                mnth = sheetName[currmonth].value
                if ((sheetName[currmonth].value is None) or (len(sheetName[currmonth].value) == 0)):
                    self.__printFalg and print("Unable to parse string currmonth", sheetName[currmonth].value)
                    mnth = sheetName[currmonth].value
                else:
                    mnth = sheetName[currmonth].value.split("~")[0].strip().split("/")[1:]

                monthAndYear = '/'.join(mnth).strip()
                for datasets in range(0, len(cellcolrange)):
                    del puchtime[:]
                    #                 print("*" * 80)
                    for rws in range(cellrwrangest, sheetName.max_row):
                        # get the cell start as per config file and append the number of rows to select the range
                        stCell = cellcolrange[datasets][0] + str(rws)
                        endCell = cellcolrange[datasets][1] + str(rws)
                        self.__printFalg and print("getSheetData", sheetName, "Cell numbers", cellcolrange[datasets][0] + str(rws),
                                                   cellcolrange[datasets][1] + str(rws))
                        for rowOfCellObjects in sheetName[stCell:endCell]:
                            if ((rowOfCellObjects[0].value is None) or (len(rowOfCellObjects[0].value) == 0)):
                                print("Unable to parse string day of week", rowOfCellObjects[0].value)
                                dayNumber = 100
                                dayOfWeek = 100
                            else:
                                dayNumber = rowOfCellObjects[0].value.split(" ")[0].strip()
                                dayOfWeek = rowOfCellObjects[0].value.split(" ")[1].strip()
                                dateComponent = dayNumber + "/" + monthAndYear

                                self.__printFalg and print("getSheetData ", rowOfCellObjects[0].value, rowOfCellObjects[1].value,
                                                           rowOfCellObjects[3].value)
                                # inTime
                                if rowOfCellObjects[1].value is not None:
                                    # cleanInTime = datetime.datetime.strptime(self.massageData(rowOfCellObjects[1].value), '%H:%M:%S').time()
                                    # inTime = dateComponent + " " + self.massageData(rowOfCellObjects[1].value).strftime("%H:%M:%S").strip()
                                    inTime = dateComponent + " " + rowOfCellObjects[1].value.strftime("%H:%M:%S").strip()
                                    # inTime = dateComponent + " " + self.massageData(rowOfCellObjects[1].value).strip()
                                else:
                                    inTime = dateComponent + " " + datetime.time(0, 0, 0).strftime("%H:%M:%S")

                                # OutTime
                                if rowOfCellObjects[3].value is not None:
                                    # cleanOutTime = datetime.datetime.strptime(self.massageData(rowOfCellObjects[3].value), '%H:%M:%S').time()
                                    # outTime = dateComponent + " " + self.massageData(rowOfCellObjects[3].value).strftime("%H:%M:%S").strip()
                                    outTime = dateComponent + " " + rowOfCellObjects[3].value.strftime("%H:%M:%S").strip()
                                    # inTime = dateComponent + " " + self.massageData(rowOfCellObjects[3].value).strip()
                                else:
                                    outTime = dateComponent + " " + datetime.time(0, 0, 0).strftime("%H:%M:%S")

                                dayOfWeekParsed = self.parseDay(dayOfWeek)

                                # print(dayOfWeekParsed, inTime, outTime, type(cleanInTime),type(cleanOutTime) )

                                puchtime.append({
                                    "day": parse(dateComponent),
                                    "startTime": parse(inTime),
                                    "endTime": parse(outTime),
                                    "dayOfWeek": dayOfWeekParsed
                                })

                    extractedData[sheetName[namelist[datasets]].value.replace('.', '_')] = {
                        "sheet": worksheets[sheet],
                        # "month" : sheetName[currmonth].value ,
                        "puchtime": copy.copy(puchtime),
                        "dayAbsence": sheetName[daysAbsent[datasets]].value,
                        "daysWorked": sheetName[daysWorked[datasets]].value
                    }
                    # })

            mnth = sheetName[currmonth].value.split("~")
            monthStart = mnth[0].strip()
            monthEnd = mnth[1].strip()
            extractedData["monthStart"] = monthStart
            extractedData["monthEnd"] = monthEnd
        except Exception as e:
            print("\n\nException getSheetData", sheetName, "Cell numbers", cellcolrange[datasets][0] + str(rws),
                                       cellcolrange[datasets][1] + str(rws))

            print("Exception getSheetData ", ":", rowOfCellObjects[0].value, ":", rowOfCellObjects[1].value, ":",
                  rowOfCellObjects[3].value, ":", type(rowOfCellObjects[1].value), ":", type(rowOfCellObjects[3].value)
                  )
            pp.pprint(str(e))
            extractedData['err'] = []
            extractedData['err'].append('Error in parsing the Excel:getSheetData' + str(e))
            raise ValueError

        return extractedData

    # get the configuration data for shop opening and closing for the particular dates
    def loadShopConfigData(self, filename):
        pp = pprint.PrettyPrinter(indent=4)
        wbObj = self.getWorkBookObj(filename)
        worksheets = self.getSheetNames(wbObj)
        cnfgjson = self.loadconfigdata()

        print("worksheets", worksheets, type(worksheets), len(worksheets))
        sheetName = self.getSheetObj(wbObj, worksheets[0])
        print("sheetName", sheetName, " sheetName.max_row", sheetName.max_row)
        shopConfigData = {}
        empDetailsdata = {}
        # shop configurationd details related to shop holidays, special start times
        if worksheets[0] == "ShopConfig":
            colRange = cnfgjson["ShopConfigExcel"]["colRange"]
            rowRange = cnfgjson["ShopConfigExcel"]["rowRange"]
            rwStart = int(rowRange["start"])
            rwEnd = rowRange["end"]
            print("getShopConfigData", "colRange", colRange, "rowRange", rowRange, "rwStart", rwStart, "rwEnd", rwEnd)
            for rws in range(rwStart, sheetName.max_row + 1, 1):
                stCell = colRange[0] + str(rws)
                endCell = colRange[1] + str(rws)
                print("stCell", stCell, "endCell", endCell)
                for rowOfCellObjects in sheetName[stCell:endCell]:
                    if rowOfCellObjects[0].value.strip() is not None:
                        workDate = rowOfCellObjects[0].value.strip()
                        intime = rowOfCellObjects[0].value.strip() + " " + rowOfCellObjects[1].value.strftime(
                            "%H:%M:%S").strip()
                        outtime = rowOfCellObjects[0].value.strip() + " " + rowOfCellObjects[2].value.strftime(
                            "%H:%M:%S").strip()
                        entryType = rowOfCellObjects[3].value
                        comments = rowOfCellObjects[4].value
                        print("dy", workDate, "intime", intime, "outtime", outtime, "entryType", entryType, "comments",
                              comments)
                        shopConfigData[workDate + ":" + entryType] = {"workdate": parse(workDate),
                                                                      "intime": parse(intime),
                                                                      "outtime": parse(outtime),
                                                                      "entryType": entryType, "comments": comments}

        # Employees details related to sal and shift timings
        sheetName = self.getSheetObj(wbObj, worksheets[1])

        if worksheets[1] == "EmployeeInfo":
            colRange = cnfgjson["employeeDetails"]["colRange"]
            rowRange = cnfgjson["employeeDetails"]["rowRange"]
            rwStart = int(rowRange["start"])
            rwEnd = rowRange["end"]
            print("getShopConfigData", "colRange", colRange, "rowRange", rowRange, "rwStart", rwStart, "rwEnd", rwEnd)
            for rws in range(rwStart, sheetName.max_row + 1, 1):
                stCell = colRange[0] + str(rws)
                endCell = colRange[1] + str(rws)
                print("stCell", stCell, "endCell", endCell)
                for rowOfCellObjects in sheetName[stCell:endCell]:
                    if rowOfCellObjects[0].value.strip() is not None:
                        empName = rowOfCellObjects[0].value.replace('.', '_').strip()  ##replace . in name with _
                        shift = rowOfCellObjects[1].value.strip()
                        salary = rowOfCellObjects[2].value
                        print("empName", empName, "shift", shift, "salary", salary)
                        empDetailsdata[empName] = {"shift": shift, "salary": salary}

        return shopConfigData, empDetailsdata

    def applyComputeSalaryLogic(self, empobj):

        # salary calculations
        # Variables to hold the salary calculations
        for dailystats in empobj["workhours"]:
            if dailystats["DailyDeductions"] > 0:
                empobj["totaldeductions"] = round(empobj["totaldeductions"] + dailystats["DailyDeductions"], 2)
            if dailystats["DailyAdditions"] > 0:
                empobj["totalAdditions"] = round(empobj["totalAdditions"] + dailystats["DailyAdditions"], 2)

        ## if absent on sunday then reduce salary by 2 days
        if empobj["totalSundaysAbsent"] > 0:
            empobj["totaldeductions"] = round( empobj["totaldeductions"] + empobj["dailySalary"] * empobj["totalSundaysAbsent"] * 2 ,2)

        ## Check holiday surplus or deficiency
        ## if holiday taken are more than allowed then reduce the salary by extra days taken as holiday
        ## employee is in deficit, update totaldeductions
        if empobj["totalHolidaytaken"] > empobj["totalHolidaysAllowed"]:
            # update deduction as daily salary into total extra days consumed
            empobj["totaldeductions"] = round ( empobj["totaldeductions"] + empobj["dailySalary"] * (
                        empobj["totalHolidaytaken"] - empobj["totalHolidaysAllowed"]),2)

        ## employee is in surplus, update totalAdditions
        if empobj["totalHolidaytaken"] < empobj["totalHolidaysAllowed"]:
            # update totalAdditions as daily salary into total extra days not used
            empobj["totalAdditions"] = round(  empobj["totalAdditions"] + (empobj["dailySalary"] * (
                        empobj["totalHolidaysAllowed"] - empobj["totalHolidaytaken"]))
                                                ,2)

        # if employee worked on a company defined holiday
        if empobj["totalHolidaysWorked"] > 0:
            empobj["totalAdditions"] = round( empobj["totalAdditions"] + (empobj["dailySalary"] * empobj["totalHolidaysWorked"]),2)

        empobj['CalculatedSalary'] = round(empobj["totalMonthlySalary"] - empobj["totaldeductions"] + empobj["totalAdditions"],2)
        return empobj

    # apply the business rules for you and i based on the configuration sepcified in the shop config
    def applyBuinessRules(self, excelEmpData):
        print("Inside applyBuinessRules")
        pp = pprint.PrettyPrinter(indent=4)
        excelEmployeeParsedData = excelEmpData
        # emp json
        # excelEmployeeParsedData = {
        #     'govind': {'dayAbsence': 4,
        #                'daysWorked': 26,
        #                'puchtime': [{'day': datetime.datetime(2016, 1, 11, 0, 0),
        #                              'dayOfWeek': 'Tue',
        #                              'endTime': datetime.datetime(2016, 1, 11, 0, 0),
        #                              'startTime': datetime.datetime(2016, 1, 11, 0, 0)},
        #                             {'day': datetime.datetime(2016, 2, 11, 0, 0),
        #                              'dayOfWeek': 'Wed',
        #                              'endTime': datetime.datetime(2016, 2, 11, 20, 39, 59),
        #                              'startTime': datetime.datetime(2016, 2, 11, 11, 1, 59)},
        #                             {'day': datetime.datetime(2016, 3, 11, 0, 0),
        #                              'dayOfWeek': 'Thu',
        #                              'endTime': datetime.datetime(2016, 3, 11, 20, 33, 59),
        #                              'startTime': datetime.datetime(2016, 3, 11, 10, 19, 59)},
        #                             {'day': datetime.datetime(2016, 4, 11, 0, 0),
        #                              'dayOfWeek': 'Fri',
        #                              'endTime': datetime.datetime(2016, 4, 11, 20, 36, 59),
        #                              'startTime': datetime.datetime(2016, 4, 11, 10, 32, 59)},
        #                             {'day': datetime.datetime(2016, 5, 11, 0, 0),
        #                              'dayOfWeek': 'Sat',
        #                              'endTime': datetime.datetime(2016, 5, 11, 20, 34, 59),
        #                              'startTime': datetime.datetime(2016, 5, 11, 10, 53, 59)},
        #                             {'day': datetime.datetime(2016, 6, 11, 0, 0),
        #                              'dayOfWeek': 'Sun',
        #                              'endTime': datetime.datetime(2016, 6, 11, 20, 47, 59),
        #                              'startTime': datetime.datetime(2016, 6, 11, 10, 28, 59)},
        #                             {'day': datetime.datetime(2016, 7, 11, 0, 0),
        #                              'dayOfWeek': 'Mon',
        #                              'endTime': datetime.datetime(2016, 7, 11, 20, 30, 59),
        #                              'startTime': datetime.datetime(2016, 7, 11, 10, 24, 59)},
        #                             {'day': datetime.datetime(2016, 8, 11, 0, 0),
        #                              'dayOfWeek': 'Tue',
        #                              'endTime': datetime.datetime(2016, 8, 11, 20, 34, 59),
        #                              'startTime': datetime.datetime(2016, 8, 11, 10, 23, 59)},
        #                             {'day': datetime.datetime(2016, 9, 11, 0, 0),
        #                              'dayOfWeek': 'Wed',
        #                              'endTime': datetime.datetime(2016, 9, 11, 20, 6, 59),
        #                              'startTime': datetime.datetime(2016, 9, 11, 11, 10, 59)},
        #                             {'day': datetime.datetime(2016, 10, 11, 0, 0),
        #                              'dayOfWeek': 'Thu',
        #                              'endTime': datetime.datetime(2016, 10, 11, 0, 0),
        #                              'startTime': datetime.datetime(2016, 10, 11, 20, 34, 59)},
        #                             {'day': datetime.datetime(2016, 11, 11, 0, 0),
        #                              'dayOfWeek': 'Fri',
        #                              'endTime': datetime.datetime(2016, 11, 11, 20, 38, 59),
        #                              'startTime': datetime.datetime(2016, 11, 11, 10, 33, 59)},
        #                             {'day': datetime.datetime(2016, 12, 11, 0, 0),
        #                              'dayOfWeek': 'Sat',
        #                              'endTime': datetime.datetime(2016, 12, 11, 0, 0),
        #                              'startTime': datetime.datetime(2016, 12, 11, 0, 0)},
        #                             {'day': datetime.datetime(2016, 11, 13, 0, 0),
        #                              'dayOfWeek': 'Sun',
        #                              'endTime': datetime.datetime(2016, 11, 13, 20, 31, 59),
        #                              'startTime': datetime.datetime(2016, 11, 13, 10, 49, 59)},
        #                             {'day': datetime.datetime(2016, 11, 14, 0, 0),
        #                              'dayOfWeek': 'Mon',
        #                              'endTime': datetime.datetime(2016, 11, 14, 20, 28, 59),
        #                              'startTime': datetime.datetime(2016, 11, 14, 10, 28, 59)},
        #                             {'day': datetime.datetime(2016, 11, 15, 0, 0),
        #                              'dayOfWeek': 'Tue',
        #                              'endTime': datetime.datetime(2016, 11, 15, 20, 44, 59),
        #                              'startTime': datetime.datetime(2016, 11, 15, 10, 32, 59)},
        #                             {'day': datetime.datetime(2016, 11, 16, 0, 0),
        #                              'dayOfWeek': 'Wed',
        #                              'endTime': datetime.datetime(2016, 11, 16, 20, 34, 59),
        #                              'startTime': datetime.datetime(2016, 11, 16, 10, 27, 59)},
        #                             {'day': datetime.datetime(2016, 11, 17, 0, 0),
        #                              'dayOfWeek': 'Thu',
        #                              'endTime': datetime.datetime(2016, 11, 17, 20, 39, 59),
        #                              'startTime': datetime.datetime(2016, 11, 17, 12, 37, 59)},
        #                             {'day': datetime.datetime(2016, 11, 18, 0, 0),
        #                              'dayOfWeek': 'Fri',
        #                              'endTime': datetime.datetime(2016, 11, 18, 0, 0),
        #                              'startTime': datetime.datetime(2016, 11, 18, 0, 0)},
        #                             {'day': datetime.datetime(2016, 11, 19, 0, 0),
        #                              'dayOfWeek': 'Sat',
        #                              'endTime': datetime.datetime(2016, 11, 19, 20, 25, 59),
        #                              'startTime': datetime.datetime(2016, 11, 19, 10, 36, 59)},
        #                             {'day': datetime.datetime(2016, 11, 20, 0, 0),
        #                              'dayOfWeek': 'Sun',
        #                              'endTime': datetime.datetime(2016, 11, 20, 19, 31, 59),
        #                              'startTime': datetime.datetime(2016, 11, 20, 10, 38, 59)},
        #                             {'day': datetime.datetime(2016, 11, 21, 0, 0),
        #                              'dayOfWeek': 'Mon',
        #                              'endTime': datetime.datetime(2016, 11, 21, 20, 27, 59),
        #                              'startTime': datetime.datetime(2016, 11, 21, 14, 5, 59)},
        #                             {'day': datetime.datetime(2016, 11, 22, 0, 0),
        #                              'dayOfWeek': 'Tue',
        #                              'endTime': datetime.datetime(2016, 11, 22, 21, 6, 59),
        #                              'startTime': datetime.datetime(2016, 11, 22, 11, 40, 59)},
        #                             {'day': datetime.datetime(2016, 11, 23, 0, 0),
        #                              'dayOfWeek': 'Wed',
        #                              'endTime': datetime.datetime(2016, 11, 23, 20, 29, 59),
        #                              'startTime': datetime.datetime(2016, 11, 23, 10, 30, 59)},
        #                             {'day': datetime.datetime(2016, 11, 24, 0, 0),
        #                              'dayOfWeek': 'Thu',
        #                              'endTime': datetime.datetime(2016, 11, 24, 20, 41, 59),
        #                              'startTime': datetime.datetime(2016, 11, 24, 14, 20, 59)},
        #                             {'day': datetime.datetime(2016, 11, 25, 0, 0),
        #                              'dayOfWeek': 'Fri',
        #                              'endTime': datetime.datetime(2016, 11, 25, 20, 30, 59),
        #                              'startTime': datetime.datetime(2016, 11, 25, 10, 35, 59)},
        #                             {'day': datetime.datetime(2016, 11, 26, 0, 0),
        #                              'dayOfWeek': 'Sat',
        #                              'endTime': datetime.datetime(2016, 11, 26, 20, 38, 59),
        #                              'startTime': datetime.datetime(2016, 11, 26, 10, 32, 59)},
        #                             {'day': datetime.datetime(2016, 11, 27, 0, 0),
        #                              'dayOfWeek': 'Sun',
        #                              'endTime': datetime.datetime(2016, 11, 27, 0, 0),
        #                              'startTime': datetime.datetime(2016, 11, 27, 10, 22, 59)},
        #                             {'day': datetime.datetime(2016, 11, 28, 0, 0),
        #                              'dayOfWeek': 'Mon',
        #                              'endTime': datetime.datetime(2016, 11, 28, 0, 0),
        #                              'startTime': datetime.datetime(2016, 11, 28, 20, 23, 59)},
        #                             {'day': datetime.datetime(2016, 11, 29, 0, 0),
        #                              'dayOfWeek': 'Tue',
        #                              'endTime': "",
        #                              'startTime': datetime.datetime(2016, 11, 29, 10, 33, 59)},
        #                             {'day': datetime.datetime(2016, 11, 30, 0, 0),
        #                              'dayOfWeek': 'Wed',
        #                              'endTime': datetime.datetime(2016, 11, 30, 0, 0),
        #                              'startTime': datetime.datetime(2016, 11, 30, 0, 0)}],
        #                'sheet': '4,5,6'},
        #     # 'manoj': {   'dayAbsence': 9,
        #     #      'daysWorked': 21,
        #     #      'puchtime': [   {   'day': datetime.datetime(2016, 1, 11, 0, 0),
        #     #                          'dayOfWeek': 'Tue',
        #     #                          'endTime': datetime.datetime(2016, 1, 11, 0, 0),
        #     #                          'startTime': datetime.datetime(2016, 1, 11, 0, 0)},
        #     #                      {   'day': datetime.datetime(2016, 2, 11, 0, 0),
        #     #                          'dayOfWeek': 'Wed',
        #     #                          'endTime': datetime.datetime(2016, 2, 11, 0, 0),
        #     #                          'startTime': datetime.datetime(2016, 2, 11, 0, 0)},
        #     #                      {   'day': datetime.datetime(2016, 3, 11, 0, 0),
        #     #                          'dayOfWeek': 'Thu',
        #     #                          'endTime': datetime.datetime(2016, 3, 11, 20, 30, 59),
        #     #                          'startTime': datetime.datetime(2016, 3, 11, 10, 41, 59)},
        #     #                      {   'day': datetime.datetime(2016, 4, 11, 0, 0),
        #     #                          'dayOfWeek': 'Fri',
        #     #                          'endTime': datetime.datetime(2016, 4, 11, 20, 30, 59),
        #     #                          'startTime': datetime.datetime(2016, 4, 11, 10, 43, 59)},
        #     #                      {   'day': datetime.datetime(2016, 5, 11, 0, 0),
        #     #                          'dayOfWeek': 'Sat',
        #     #                          'endTime': datetime.datetime(2016, 5, 11, 20, 33, 59),
        #     #                          'startTime': datetime.datetime(2016, 5, 11, 10, 46, 59)},
        #     #                      {   'day': datetime.datetime(2016, 6, 11, 0, 0),
        #     #                          'dayOfWeek': 'Sun',
        #     #                          'endTime': datetime.datetime(2016, 6, 11, 20, 51, 59),
        #     #                          'startTime': datetime.datetime(2016, 6, 11, 10, 46, 59)},
        #     #                      {   'day': datetime.datetime(2016, 7, 11, 0, 0),
        #     #                          'dayOfWeek': 'Mon',
        #     #                          'endTime': datetime.datetime(2016, 7, 11, 20, 36, 59),
        #     #                          'startTime': datetime.datetime(2016, 7, 11, 11, 9, 59)},
        #     #                      {   'day': datetime.datetime(2016, 8, 11, 0, 0),
        #     #                          'dayOfWeek': 'Tue',
        #     #                          'endTime': datetime.datetime(2016, 8, 11, 20, 50, 59),
        #     #                          'startTime': datetime.datetime(2016, 8, 11, 11, 3, 59)},
        #     #                      {   'day': datetime.datetime(2016, 9, 11, 0, 0),
        #     #                          'dayOfWeek': 'Wed',
        #     #                          'endTime': datetime.datetime(2016, 9, 11, 15, 15, 59),
        #     #                          'startTime': datetime.datetime(2016, 9, 11, 11, 19, 59)},
        #     #                      {   'day': datetime.datetime(2016, 10, 11, 0, 0),
        #     #                          'dayOfWeek': 'Thu',
        #     #                          'endTime': datetime.datetime(2016, 10, 11, 0, 0),
        #     #                          'startTime': datetime.datetime(2016, 10, 11, 0, 0)},
        #     #                      {   'day': datetime.datetime(2016, 11, 11, 0, 0),
        #     #                          'dayOfWeek': 'Fri',
        #     #                          'endTime': datetime.datetime(2016, 11, 11, 0, 0),
        #     #                          'startTime': datetime.datetime(2016, 11, 11, 0, 0)},
        #     #                      {   'day': datetime.datetime(2016, 12, 11, 0, 0),
        #     #                          'dayOfWeek': 'Sat',
        #     #                          'endTime': datetime.datetime(2016, 12, 11, 0, 0),
        #     #                          'startTime': datetime.datetime(2016, 12, 11, 0, 0)},
        #     #                      {   'day': datetime.datetime(2016, 11, 13, 0, 0),
        #     #                          'dayOfWeek': 'Sun',
        #     #                          'endTime': datetime.datetime(2016, 11, 13, 20, 36, 59),
        #     #                          'startTime': datetime.datetime(2016, 11, 13, 11, 2, 59)},
        #     #                      {   'day': datetime.datetime(2016, 11, 14, 0, 0),
        #     #                          'dayOfWeek': 'Mon',
        #     #                          'endTime': datetime.datetime(2016, 11, 14, 21, 0),
        #     #                          'startTime': datetime.datetime(2016, 11, 14, 10, 35, 59)},
        #     #                      {   'day': datetime.datetime(2016, 11, 15, 0, 0),
        #     #                          'dayOfWeek': 'Tue',
        #     #                          'endTime': datetime.datetime(2016, 11, 15, 20, 32, 59),
        #     #                          'startTime': datetime.datetime(2016, 11, 15, 10, 35, 59)},
        #     #                      {   'day': datetime.datetime(2016, 11, 16, 0, 0),
        #     #                          'dayOfWeek': 'Wed',
        #     #                          'endTime': datetime.datetime(2016, 11, 16, 20, 25, 59),
        #     #                          'startTime': datetime.datetime(2016, 11, 16, 11, 9, 59)},
        #     #                      {   'day': datetime.datetime(2016, 11, 17, 0, 0),
        #     #                          'dayOfWeek': 'Thu',
        #     #                          'endTime': datetime.datetime(2016, 11, 17, 20, 30, 59),
        #     #                          'startTime': datetime.datetime(2016, 11, 17, 11, 32, 59)},
        #     #                      {   'day': datetime.datetime(2016, 11, 18, 0, 0),
        #     #                          'dayOfWeek': 'Fri',
        #     #                          'endTime': datetime.datetime(2016, 11, 18, 20, 13, 59),
        #     #                          'startTime': datetime.datetime(2016, 11, 18, 11, 1, 59)},
        #     #                      {   'day': datetime.datetime(2016, 11, 19, 0, 0),
        #     #                          'dayOfWeek': 'Sat',
        #     #                          'endTime': datetime.datetime(2016, 11, 19, 0, 0),
        #     #                          'startTime': datetime.datetime(2016, 11, 19, 0, 0)},
        #     #                      {   'day': datetime.datetime(2016, 11, 20, 0, 0),
        #     #                          'dayOfWeek': 'Sun',
        #     #                          'endTime': datetime.datetime(2016, 11, 20, 21, 4, 59),
        #     #                          'startTime': datetime.datetime(2016, 11, 20, 10, 36, 59)},
        #     #                      {   'day': datetime.datetime(2016, 11, 21, 0, 0),
        #     #                          'dayOfWeek': 'Mon',
        #     #                          'endTime': datetime.datetime(2016, 11, 21, 20, 48, 59),
        #     #                          'startTime': datetime.datetime(2016, 11, 21, 10, 55, 59)},
        #     #                      {   'day': datetime.datetime(2016, 11, 22, 0, 0),
        #     #                          'dayOfWeek': 'Tue',
        #     #                          'endTime': datetime.datetime(2016, 11, 22, 20, 33, 59),
        #     #                          'startTime': datetime.datetime(2016, 11, 22, 10, 34, 59)},
        #     #                      {   'day': datetime.datetime(2016, 11, 23, 0, 0),
        #     #                          'dayOfWeek': 'Wed',
        #     #                          'endTime': datetime.datetime(2016, 11, 23, 20, 32, 59),
        #     #                          'startTime': datetime.datetime(2016, 11, 23, 10, 51, 59)},
        #     #                      {   'day': datetime.datetime(2016, 11, 24, 0, 0),
        #     #                          'dayOfWeek': 'Thu',
        #     #                          'endTime': datetime.datetime(2016, 11, 24, 20, 31, 59),
        #     #                          'startTime': datetime.datetime(2016, 11, 24, 11, 21, 59)},
        #     #                      {   'day': datetime.datetime(2016, 11, 25, 0, 0),
        #     #                          'dayOfWeek': 'Fri',
        #     #                          'endTime': datetime.datetime(2016, 11, 25, 20, 46, 59),
        #     #                          'startTime': datetime.datetime(2016, 11, 25, 11, 18, 59)},
        #     #                      {   'day': datetime.datetime(2016, 11, 26, 0, 0),
        #     #                          'dayOfWeek': 'Sat',
        #     #                          'endTime': datetime.datetime(2016, 11, 26, 20, 25, 59),
        #     #                          'startTime': datetime.datetime(2016, 11, 26, 11, 28, 59)},
        #     #                      {   'day': datetime.datetime(2016, 11, 27, 0, 0),
        #     #                          'dayOfWeek': 'Sun',
        #     #                          'endTime': datetime.datetime(2016, 11, 27, 0, 0),
        #     #                          'startTime': datetime.datetime(2016, 11, 27, 0, 0)},
        #     #                      {   'day': datetime.datetime(2016, 11, 28, 0, 0),
        #     #                          'dayOfWeek': 'Mon',
        #     #                          'endTime': datetime.datetime(2016, 11, 28, 0, 0),
        #     #                          'startTime': datetime.datetime(2016, 11, 28, 0, 0)},
        #     #                      {   'day': datetime.datetime(2016, 11, 29, 0, 0),
        #     #                          'dayOfWeek': 'Tue',
        #     #                          'endTime': datetime.datetime(2016, 11, 29, 0, 0),
        #     #                          'startTime': datetime.datetime(2016, 11, 29, 0, 0)},
        #     #                      {   'day': datetime.datetime(2016, 11, 30, 0, 0),
        #     #                          'dayOfWeek': 'Wed',
        #     #                          'endTime': datetime.datetime(2016, 11, 30, 20, 55, 59),
        #     #                          'startTime': datetime.datetime(2016, 11, 30, 10, 55, 59)}],
        #     #      'sheet': '1,2,3'},
        #     'monthEnd': '30/11/2016',
        #     'monthStart': '01/11/2016',
        #     'naveen': {'dayAbsence': 2,
        #                'daysWorked': 28,
        #                'puchtime': [{'day': datetime.datetime(2016, 1, 11, 0, 0),
        #                              'dayOfWeek': 'Tue',
        #                              'endTime': datetime.datetime(2016, 1, 11, 20, 32, 59),
        #                              'startTime': datetime.datetime(2016, 1, 11, 10, 34, 59)},
        #                             {'day': datetime.datetime(2016, 2, 11, 0, 0),
        #                              'dayOfWeek': 'Wed',
        #                              'endTime': datetime.datetime(2016, 2, 11, 20, 43, 59),
        #                              'startTime': datetime.datetime(2016, 2, 11, 10, 41, 59)},
        #                             {'day': datetime.datetime(2016, 3, 11, 0, 0),
        #                              'dayOfWeek': 'Thu',
        #                              'endTime': datetime.datetime(2016, 3, 11, 20, 32, 59),
        #                              'startTime': datetime.datetime(2016, 3, 11, 10, 22, 59)},
        #                             {'day': datetime.datetime(2016, 4, 11, 0, 0),
        #                              'dayOfWeek': 'Fri',
        #                              'endTime': datetime.datetime(2016, 4, 11, 20, 36, 59),
        #                              'startTime': datetime.datetime(2016, 4, 11, 10, 31, 59)},
        #                             {'day': datetime.datetime(2016, 5, 11, 0, 0),
        #                              'dayOfWeek': 'Sat',
        #                              'endTime': datetime.datetime(2016, 5, 11, 14, 10, 59),
        #                              'startTime': datetime.datetime(2016, 5, 11, 10, 30, 59)},
        #                             {'day': datetime.datetime(2016, 6, 11, 0, 0),
        #                              'dayOfWeek': 'Sun',
        #                              'endTime': datetime.datetime(2016, 6, 11, 20, 51, 59),
        #                              'startTime': datetime.datetime(2016, 6, 11, 10, 58, 59)},
        #                             {'day': datetime.datetime(2016, 7, 11, 0, 0),
        #                              'dayOfWeek': 'Mon',
        #                              'endTime': datetime.datetime(2016, 7, 11, 0, 0),
        #                              'startTime': datetime.datetime(2016, 7, 11, 10, 30, 59)},
        #                             {'day': datetime.datetime(2016, 8, 11, 0, 0),
        #                              'dayOfWeek': 'Tue',
        #                              'endTime': datetime.datetime(2016, 8, 11, 20, 34, 59),
        #                              'startTime': datetime.datetime(2016, 8, 11, 10, 23, 59)},
        #                             {'day': datetime.datetime(2016, 9, 11, 0, 0),
        #                              'dayOfWeek': 'Wed',
        #                              'endTime': datetime.datetime(2016, 9, 11, 20, 37, 59),
        #                              'startTime': datetime.datetime(2016, 9, 11, 10, 34, 59)},
        #                             {'day': datetime.datetime(2016, 10, 11, 0, 0),
        #                              'dayOfWeek': 'Thu',
        #                              'endTime': datetime.datetime(2016, 10, 11, 20, 37, 59),
        #                              'startTime': datetime.datetime(2016, 10, 11, 10, 40, 59)},
        #                             {'day': datetime.datetime(2016, 11, 11, 0, 0),
        #                              'dayOfWeek': 'Fri',
        #                              'endTime': datetime.datetime(2016, 11, 11, 0, 0),
        #                              'startTime': datetime.datetime(2016, 11, 11, 0, 0)},
        #                             {'day': datetime.datetime(2016, 12, 11, 0, 0),
        #                              'dayOfWeek': 'Sat',
        #                              'endTime': datetime.datetime(2016, 12, 11, 20, 29, 59),
        #                              'startTime': datetime.datetime(2016, 12, 11, 10, 35, 59)},
        #                             {'day': datetime.datetime(2016, 11, 13, 0, 0),
        #                              'dayOfWeek': 'Sun',
        #                              'endTime': datetime.datetime(2016, 11, 13, 20, 38, 59),
        #                              'startTime': datetime.datetime(2016, 11, 13, 11, 28, 59)},
        #                             {'day': datetime.datetime(2016, 11, 14, 0, 0),
        #                              'dayOfWeek': 'Mon',
        #                              'endTime': datetime.datetime(2016, 11, 14, 20, 29, 59),
        #                              'startTime': datetime.datetime(2016, 11, 14, 10, 30)},
        #                             {'day': datetime.datetime(2016, 11, 15, 0, 0),
        #                              'dayOfWeek': 'Tue',
        #                              'endTime': datetime.datetime(2016, 11, 15, 20, 45, 59),
        #                              'startTime': datetime.datetime(2016, 11, 15, 10, 32, 59)},
        #                             {'day': datetime.datetime(2016, 11, 16, 0, 0),
        #                              'dayOfWeek': 'Wed',
        #                              'endTime': datetime.datetime(2016, 11, 16, 0, 0),
        #                              'startTime': datetime.datetime(2016, 11, 16, 0, 0)},
        #                             {'day': datetime.datetime(2016, 11, 17, 0, 0),
        #                              'dayOfWeek': 'Thu',
        #                              'endTime': datetime.datetime(2016, 11, 17, 20, 38, 59),
        #                              'startTime': datetime.datetime(2016, 11, 17, 10, 27, 59)},
        #                             {'day': datetime.datetime(2016, 11, 18, 0, 0),
        #                              'dayOfWeek': 'Fri',
        #                              'endTime': datetime.datetime(2016, 11, 18, 20, 33, 59),
        #                              'startTime': datetime.datetime(2016, 11, 18, 10, 33, 59)},
        #                             {'day': datetime.datetime(2016, 11, 19, 0, 0),
        #                              'dayOfWeek': 'Sat',
        #                              'endTime': datetime.datetime(2016, 11, 19, 20, 27, 59),
        #                              'startTime': datetime.datetime(2016, 11, 19, 10, 30, 59)},
        #                             {'day': datetime.datetime(2016, 11, 20, 0, 0),
        #                              'dayOfWeek': 'Sun',
        #                              'endTime': datetime.datetime(2016, 11, 20, 0, 0),
        #                              'startTime': datetime.datetime(2016, 11, 20, 20, 45, 59)},
        #                             {'day': datetime.datetime(2016, 11, 21, 0, 0),
        #                              'dayOfWeek': 'Mon',
        #                              'endTime': datetime.datetime(2016, 11, 21, 20, 27, 59),
        #                              'startTime': datetime.datetime(2016, 11, 21, 10, 28, 59)},
        #                             {'day': datetime.datetime(2016, 11, 22, 0, 0),
        #                              'dayOfWeek': 'Tue',
        #                              'endTime': datetime.datetime(2016, 11, 22, 14, 13, 59),
        #                              'startTime': datetime.datetime(2016, 11, 22, 10, 34, 59)},
        #                             {'day': datetime.datetime(2016, 11, 23, 0, 0),
        #                              'dayOfWeek': 'Wed',
        #                              'endTime': datetime.datetime(2016, 11, 23, 0, 0),
        #                              'startTime': datetime.datetime(2016, 11, 23, 20, 30, 59)},
        #                             {'day': datetime.datetime(2016, 11, 24, 0, 0),
        #                              'dayOfWeek': 'Thu',
        #                              'endTime': datetime.datetime(2016, 11, 24, 20, 41, 59),
        #                              'startTime': datetime.datetime(2016, 11, 24, 10, 46, 59)},
        #                             {'day': datetime.datetime(2016, 11, 25, 0, 0),
        #                              'dayOfWeek': 'Fri',
        #                              'endTime': datetime.datetime(2016, 11, 25, 20, 32, 59),
        #                              'startTime': datetime.datetime(2016, 11, 25, 10, 28, 59)},
        #                             {'day': datetime.datetime(2016, 11, 26, 0, 0),
        #                              'dayOfWeek': 'Sat',
        #                              'endTime': datetime.datetime(2016, 11, 26, 20, 38, 59),
        #                              'startTime': datetime.datetime(2016, 11, 26, 10, 56, 59)},
        #                             {'day': datetime.datetime(2016, 11, 27, 0, 0),
        #                              'dayOfWeek': 'Sun',
        #                              'endTime': datetime.datetime(2016, 11, 27, 0, 0),
        #                              'startTime': datetime.datetime(2016, 11, 27, 11, 39, 59)},
        #                             {'day': datetime.datetime(2016, 11, 28, 0, 0),
        #                              'dayOfWeek': 'Mon',
        #                              'endTime': datetime.datetime(2016, 11, 28, 20, 25, 59),
        #                              'startTime': datetime.datetime(2016, 11, 28, 10, 40, 59)},
        #                             {'day': datetime.datetime(2016, 11, 29, 0, 0),
        #                              'dayOfWeek': 'Tue',
        #                              'endTime': datetime.datetime(2016, 11, 29, 20, 29, 59),
        #                              'startTime': datetime.datetime(2016, 11, 29, 10, 28, 59)},
        #                             {'day': datetime.datetime(2016, 11, 30, 0, 0),
        #                              'dayOfWeek': 'Wed',
        #                              'endTime': datetime.datetime(2016, 11, 30, 0, 0),
        #                              'startTime': datetime.datetime(2016, 11, 30, 10, 30)}],
        #                'sheet': '4,5,6'},
        #     'prem': {'dayAbsence': 7,
        #              'daysWorked': 23,
        #              'puchtime': [{'day': datetime.datetime(2016, 1, 11, 0, 0),
        #                            'dayOfWeek': 'Tue',
        #                            'endTime': datetime.datetime(2016, 1, 11, 20, 40, 59),
        #                            'startTime': datetime.datetime(2016, 1, 11, 10, 32, 59)},
        #                           {'day': datetime.datetime(2016, 2, 11, 0, 0),
        #                            'dayOfWeek': 'Wed',
        #                            'endTime': datetime.datetime(2016, 2, 11, 20, 35, 59),
        #                            'startTime': datetime.datetime(2016, 2, 11, 10, 25, 59)},
        #                           {'day': datetime.datetime(2016, 3, 11, 0, 0),
        #                            'dayOfWeek': 'Thu',
        #                            'endTime': datetime.datetime(2016, 3, 11, 20, 29, 59),
        #                            'startTime': datetime.datetime(2016, 3, 11, 10, 17, 59)},
        #                           {'day': datetime.datetime(2016, 4, 11, 0, 0),
        #                            'dayOfWeek': 'Fri',
        #                            'endTime': datetime.datetime(2016, 4, 11, 0, 0),
        #                            'startTime': datetime.datetime(2016, 4, 11, 0, 0)},
        #                           {'day': datetime.datetime(2016, 5, 11, 0, 0),
        #                            'dayOfWeek': 'Sat',
        #                            'endTime': datetime.datetime(2016, 5, 11, 0, 0),
        #                            'startTime': datetime.datetime(2016, 5, 11, 0, 0)},
        #                           {'day': datetime.datetime(2016, 6, 11, 0, 0),
        #                            'dayOfWeek': 'Sun',
        #                            'endTime': datetime.datetime(2016, 6, 11, 0, 0),
        #                            'startTime': datetime.datetime(2016, 6, 11, 0, 0)},
        #                           {'day': datetime.datetime(2016, 7, 11, 0, 0),
        #                            'dayOfWeek': 'Mon',
        #                            'endTime': datetime.datetime(2016, 7, 11, 0, 0),
        #                            'startTime': datetime.datetime(2016, 7, 11, 0, 0)},
        #                           {'day': datetime.datetime(2016, 8, 11, 0, 0),
        #                            'dayOfWeek': 'Tue',
        #                            'endTime': datetime.datetime(2016, 8, 11, 0, 0),
        #                            'startTime': datetime.datetime(2016, 8, 11, 0, 0)},
        #                           {'day': datetime.datetime(2016, 9, 11, 0, 0),
        #                            'dayOfWeek': 'Wed',
        #                            'endTime': datetime.datetime(2016, 9, 11, 20, 31, 59),
        #                            'startTime': datetime.datetime(2016, 9, 11, 10, 26, 59)},
        #                           {'day': datetime.datetime(2016, 10, 11, 0, 0),
        #                            'dayOfWeek': 'Thu',
        #                            'endTime': datetime.datetime(2016, 10, 11, 0, 0),
        #                            'startTime': datetime.datetime(2016, 10, 11, 0, 0)},
        #                           {'day': datetime.datetime(2016, 11, 11, 0, 0),
        #                            'dayOfWeek': 'Fri',
        #                            'endTime': datetime.datetime(2016, 11, 11, 20, 31, 59),
        #                            'startTime': datetime.datetime(2016, 11, 11, 10, 30)},
        #                           {'day': datetime.datetime(2016, 12, 11, 0, 0),
        #                            'dayOfWeek': 'Sat',
        #                            'endTime': datetime.datetime(2016, 12, 11, 20, 30, 59),
        #                            'startTime': datetime.datetime(2016, 12, 11, 11, 29, 59)},
        #                           {'day': datetime.datetime(2016, 11, 13, 0, 0),
        #                            'dayOfWeek': 'Sun',
        #                            'endTime': datetime.datetime(2016, 11, 13, 19, 59, 59),
        #                            'startTime': datetime.datetime(2016, 11, 13, 10, 54, 59)},
        #                           {'day': datetime.datetime(2016, 11, 14, 0, 0),
        #                            'dayOfWeek': 'Mon',
        #                            'endTime': datetime.datetime(2016, 11, 14, 20, 30, 59),
        #                            'startTime': datetime.datetime(2016, 11, 14, 10, 27, 59)},
        #                           {'day': datetime.datetime(2016, 11, 15, 0, 0),
        #                            'dayOfWeek': 'Tue',
        #                            'endTime': datetime.datetime(2016, 11, 15, 20, 30, 59),
        #                            'startTime': datetime.datetime(2016, 11, 15, 10, 23, 59)},
        #                           {'day': datetime.datetime(2016, 11, 16, 0, 0),
        #                            'dayOfWeek': 'Wed',
        #                            'endTime': datetime.datetime(2016, 11, 16, 20, 29, 59),
        #                            'startTime': datetime.datetime(2016, 11, 16, 10, 28, 59)},
        #                           {'day': datetime.datetime(2016, 11, 17, 0, 0),
        #                            'dayOfWeek': 'Thu',
        #                            'endTime': datetime.datetime(2016, 11, 17, 20, 32, 59),
        #                            'startTime': datetime.datetime(2016, 11, 17, 10, 27, 59)},
        #                           {'day': datetime.datetime(2016, 11, 18, 0, 0),
        #                            'dayOfWeek': 'Fri',
        #                            'endTime': datetime.datetime(2016, 11, 18, 20, 30, 59),
        #                            'startTime': datetime.datetime(2016, 11, 18, 10, 33, 59)},
        #                           {'day': datetime.datetime(2016, 11, 19, 0, 0),
        #                            'dayOfWeek': 'Sat',
        #                            'endTime': datetime.datetime(2016, 11, 19, 0, 0),
        #                            'startTime': datetime.datetime(2016, 11, 19, 0, 0)},
        #                           {'day': datetime.datetime(2016, 11, 20, 0, 0),
        #                            'dayOfWeek': 'Sun',
        #                            'endTime': datetime.datetime(2016, 11, 20, 20, 36, 59),
        #                            'startTime': datetime.datetime(2016, 11, 20, 10, 36, 59)},
        #                           {'day': datetime.datetime(2016, 11, 21, 0, 0),
        #                            'dayOfWeek': 'Mon',
        #                            'endTime': datetime.datetime(2016, 11, 21, 20, 36, 59),
        #                            'startTime': datetime.datetime(2016, 11, 21, 10, 32, 59)},
        #                           {'day': datetime.datetime(2016, 11, 22, 0, 0),
        #                            'dayOfWeek': 'Tue',
        #                            'endTime': datetime.datetime(2016, 11, 22, 20, 30, 59),
        #                            'startTime': datetime.datetime(2016, 11, 22, 11, 20, 59)},
        #                           {'day': datetime.datetime(2016, 11, 23, 0, 0),
        #                            'dayOfWeek': 'Wed',
        #                            'endTime': datetime.datetime(2016, 11, 23, 20, 31, 59),
        #                            'startTime': datetime.datetime(2016, 11, 23, 10, 26, 59)},
        #                           {'day': datetime.datetime(2016, 11, 24, 0, 0),
        #                            'dayOfWeek': 'Thu',
        #                            'endTime': datetime.datetime(2016, 11, 24, 20, 29, 59),
        #                            'startTime': datetime.datetime(2016, 11, 24, 10, 46, 59)},
        #                           {'day': datetime.datetime(2016, 11, 25, 0, 0),
        #                            'dayOfWeek': 'Fri',
        #                            'endTime': datetime.datetime(2016, 11, 25, 19, 6, 59),
        #                            'startTime': datetime.datetime(2016, 11, 25, 10, 28, 59)},
        #                           {'day': datetime.datetime(2016, 11, 26, 0, 0),
        #                            'dayOfWeek': 'Sat',
        #                            'endTime': datetime.datetime(2016, 11, 26, 20, 29, 59),
        #                            'startTime': datetime.datetime(2016, 11, 26, 10, 32, 59)},
        #                           {'day': datetime.datetime(2016, 11, 27, 0, 0),
        #                            'dayOfWeek': 'Sun',
        #                            'endTime': datetime.datetime(2016, 11, 27, 20, 44, 59),
        #                            'startTime': datetime.datetime(2016, 11, 27, 10, 31, 59)},
        #                           {'day': datetime.datetime(2016, 11, 28, 0, 0),
        #                            'dayOfWeek': 'Mon',
        #                            'endTime': datetime.datetime(2016, 11, 28, 20, 32, 59),
        #                            'startTime': datetime.datetime(2016, 11, 28, 10, 38, 59)},
        #                           {'day': datetime.datetime(2016, 11, 29, 0, 0),
        #                            'dayOfWeek': 'Tue',
        #                            'endTime': datetime.datetime(2016, 11, 29, 20, 54, 59),
        #                            'startTime': datetime.datetime(2016, 11, 29, 10, 24, 59)},
        #                           {'day': datetime.datetime(2016, 11, 30, 0, 0),
        #                            'dayOfWeek': 'Wed',
        #                            'endTime': datetime.datetime(2016, 11, 30, 14, 20, 59),
        #                            'startTime': datetime.datetime(2016, 11, 30, 10, 27, 59)}],
        #              'sheet': '1,2,3'},
        #     'vipin': {'dayAbsence': 2,
        #               'daysWorked': 28,
        #               'puchtime': [{'day': datetime.datetime(2016, 1, 11, 0, 0),
        #                             'dayOfWeek': 'Tue',
        #                             'endTime': datetime.datetime(2016, 1, 11, 20, 32, 59),
        #                             'startTime': datetime.datetime(2016, 1, 11, 10, 47, 59)},
        #                            {'day': datetime.datetime(2016, 2, 11, 0, 0),
        #                             'dayOfWeek': 'Wed',
        #                             'endTime': datetime.datetime(2016, 2, 11, 0, 0),
        #                             'startTime': datetime.datetime(2016, 2, 11, 0, 0)},
        #                            {'day': datetime.datetime(2016, 3, 11, 0, 0),
        #                             'dayOfWeek': 'Thu',
        #                             'endTime': datetime.datetime(2016, 3, 11, 20, 33, 59),
        #                             'startTime': datetime.datetime(2016, 3, 11, 10, 47, 59)},
        #                            {'day': datetime.datetime(2016, 4, 11, 0, 0),
        #                             'dayOfWeek': 'Fri',
        #                             'endTime': datetime.datetime(2016, 4, 11, 20, 41, 59),
        #                             'startTime': datetime.datetime(2016, 4, 11, 10, 31, 59)},
        #                            {'day': datetime.datetime(2016, 5, 11, 0, 0),
        #                             'dayOfWeek': 'Sat',
        #                             'endTime': datetime.datetime(2016, 5, 11, 20, 35, 59),
        #                             'startTime': datetime.datetime(2016, 5, 11, 12, 28, 59)},
        #                            {'day': datetime.datetime(2016, 6, 11, 0, 0),
        #                             'dayOfWeek': 'Sun',
        #                             'endTime': datetime.datetime(2016, 6, 11, 20, 33, 59),
        #                             'startTime': datetime.datetime(2016, 6, 11, 11, 23, 59)},
        #                            {'day': datetime.datetime(2016, 7, 11, 0, 0),
        #                             'dayOfWeek': 'Mon',
        #                             'endTime': datetime.datetime(2016, 7, 11, 20, 32, 59),
        #                             'startTime': datetime.datetime(2016, 7, 11, 10, 28, 59)},
        #                            {'day': datetime.datetime(2016, 8, 11, 0, 0),
        #                             'dayOfWeek': 'Tue',
        #                             'endTime': datetime.datetime(2016, 8, 11, 20, 29, 59),
        #                             'startTime': datetime.datetime(2016, 8, 11, 10, 25, 59)},
        #                            {'day': datetime.datetime(2016, 9, 11, 0, 0),
        #                             'dayOfWeek': 'Wed',
        #                             'endTime': datetime.datetime(2016, 9, 11, 20, 40, 59),
        #                             'startTime': datetime.datetime(2016, 9, 11, 10, 26, 59)},
        #                            {'day': datetime.datetime(2016, 10, 11, 0, 0),
        #                             'dayOfWeek': 'Thu',
        #                             'endTime': datetime.datetime(2016, 10, 11, 20, 33, 59),
        #                             'startTime': datetime.datetime(2016, 10, 11, 10, 25, 59)},
        #                            {'day': datetime.datetime(2016, 11, 11, 0, 0),
        #                             'dayOfWeek': 'Fri',
        #                             'endTime': datetime.datetime(2016, 11, 11, 20, 42, 59),
        #                             'startTime': datetime.datetime(2016, 11, 11, 10, 31, 59)},
        #                            {'day': datetime.datetime(2016, 12, 11, 0, 0),
        #                             'dayOfWeek': 'Sat',
        #                             'endTime': datetime.datetime(2016, 12, 11, 20, 30, 59),
        #                             'startTime': datetime.datetime(2016, 12, 11, 10, 41, 59)},
        #                            {'day': datetime.datetime(2016, 11, 13, 0, 0),
        #                             'dayOfWeek': 'Sun',
        #                             'endTime': datetime.datetime(2016, 11, 13, 19, 48, 59),
        #                             'startTime': datetime.datetime(2016, 11, 13, 10, 43, 59)},
        #                            {'day': datetime.datetime(2016, 11, 14, 0, 0),
        #                             'dayOfWeek': 'Mon',
        #                             'endTime': datetime.datetime(2016, 11, 14, 20, 34, 59),
        #                             'startTime': datetime.datetime(2016, 11, 14, 10, 46, 59)},
        #                            {'day': datetime.datetime(2016, 11, 15, 0, 0),
        #                             'dayOfWeek': 'Tue',
        #                             'endTime': datetime.datetime(2016, 11, 15, 20, 44, 59),
        #                             'startTime': datetime.datetime(2016, 11, 15, 10, 22, 59)},
        #                            {'day': datetime.datetime(2016, 11, 16, 0, 0),
        #                             'dayOfWeek': 'Wed',
        #                             'endTime': datetime.datetime(2016, 11, 16, 20, 32, 59),
        #                             'startTime': datetime.datetime(2016, 11, 16, 10, 28, 59)},
        #                            {'day': datetime.datetime(2016, 11, 17, 0, 0),
        #                             'dayOfWeek': 'Thu',
        #                             'endTime': datetime.datetime(2016, 11, 17, 21, 17, 59),
        #                             'startTime': datetime.datetime(2016, 11, 17, 11, 23, 59)},
        #                            {'day': datetime.datetime(2016, 11, 18, 0, 0),
        #                             'dayOfWeek': 'Fri',
        #                             'endTime': datetime.datetime(2016, 11, 18, 0, 0),
        #                             'startTime': datetime.datetime(2016, 11, 18, 10, 46, 59)},
        #                            {'day': datetime.datetime(2016, 11, 19, 0, 0),
        #                             'dayOfWeek': 'Sat',
        #                             'endTime': datetime.datetime(2016, 11, 19, 0, 0),
        #                             'startTime': datetime.datetime(2016, 11, 19, 0, 0)},
        #                            {'day': datetime.datetime(2016, 11, 20, 0, 0),
        #                             'dayOfWeek': 'Sun',
        #                             'endTime': datetime.datetime(2016, 11, 20, 21, 47, 59),
        #                             'startTime': datetime.datetime(2016, 11, 20, 10, 36, 59)},
        #                            {'day': datetime.datetime(2016, 11, 21, 0, 0),
        #                             'dayOfWeek': 'Mon',
        #                             'endTime': datetime.datetime(2016, 11, 21, 20, 35, 59),
        #                             'startTime': datetime.datetime(2016, 11, 21, 10, 28, 59)},
        #                            {'day': datetime.datetime(2016, 11, 22, 0, 0),
        #                             'dayOfWeek': 'Tue',
        #                             'endTime': datetime.datetime(2016, 11, 22, 20, 20, 59),
        #                             'startTime': datetime.datetime(2016, 11, 22, 10, 33, 59)},
        #                            {'day': datetime.datetime(2016, 11, 23, 0, 0),
        #                             'dayOfWeek': 'Wed',
        #                             'endTime': datetime.datetime(2016, 11, 23, 20, 32, 59),
        #                             'startTime': datetime.datetime(2016, 11, 23, 10, 26, 59)},
        #                            {'day': datetime.datetime(2016, 11, 24, 0, 0),
        #                             'dayOfWeek': 'Thu',
        #                             'endTime': datetime.datetime(2016, 11, 24, 21, 0, 59),
        #                             'startTime': datetime.datetime(2016, 11, 24, 10, 46, 59)},
        #                            {'day': datetime.datetime(2016, 11, 25, 0, 0),
        #                             'dayOfWeek': 'Fri',
        #                             'endTime': datetime.datetime(2016, 11, 25, 20, 46, 59),
        #                             'startTime': datetime.datetime(2016, 11, 25, 10, 28, 59)},
        #                            {'day': datetime.datetime(2016, 11, 26, 0, 0),
        #                             'dayOfWeek': 'Sat',
        #                             'endTime': datetime.datetime(2016, 11, 26, 0, 0),
        #                             'startTime': datetime.datetime(2016, 11, 26, 10, 38, 59)},
        #                            {'day': datetime.datetime(2016, 11, 27, 0, 0),
        #                             'dayOfWeek': 'Sun',
        #                             'endTime': datetime.datetime(2016, 11, 27, 0, 0),
        #                             'startTime': datetime.datetime(2016, 11, 27, 10, 44, 59)},
        #                            {'day': datetime.datetime(2016, 11, 28, 0, 0),
        #                             'dayOfWeek': 'Mon',
        #                             'endTime': datetime.datetime(2016, 11, 28, 0, 0),
        #                             'startTime': datetime.datetime(2016, 11, 28, 11, 8, 59)},
        #                            {'day': datetime.datetime(2016, 11, 29, 0, 0),
        #                             'dayOfWeek': 'Tue',
        #                             'endTime': datetime.datetime(2016, 11, 29, 0, 0),
        #                             'startTime': datetime.datetime(2016, 11, 29, 20, 29, 59)},
        #                            {'day': datetime.datetime(2016, 11, 30, 0, 0),
        #                             'dayOfWeek': 'Wed',
        #                             'endTime': datetime.datetime(2016, 11, 30, 20, 37, 59),
        #                             'startTime': datetime.datetime(2016, 11, 30, 10, 30)}],
        #               'sheet': '1,2,3'}
        # }

        ## get the employees name, shift details and also if special dates configured from database.
        configuredDates, employeeDetails = self.getDBShopConfigData()

        # This object will be used to collect all the employees computed data
        compiledSalaries = { 'err' : [] }
        compiledSalaries["monthStart"] = excelEmployeeParsedData["monthStart"]
        compiledSalaries["monthEnd"] = excelEmployeeParsedData["monthEnd"]

        compiledSalaries["totalWorkdays"] = (datetime.datetime.strptime(excelEmployeeParsedData["monthEnd"], "%d/%m/%Y")
                                   - datetime.datetime.strptime(excelEmployeeParsedData["monthStart"],
                                                                "%d/%m/%Y")).days

        # Load configuration data and populate variables with config values
        cnfgjson = self.loadconfigdata()
        hourlyDeductionPercent = cnfgjson["salaryCalculations"]["hourlyDeductionPercent"]["val"]
        inTimeGracePeriod = cnfgjson["salaryCalculations"]["inTimeGracePeriod"]["val"]
        TotalAllowedLeavesPerMonth = cnfgjson["salaryCalculations"]["TotalLeavesPerMonth"]["val"]
        hourlyAdditionPercent = cnfgjson["salaryCalculations"]["hourlyAdditionPercent"]["val"]
        overTimeLimit = cnfgjson["salaryCalculations"]["overTimeLimit"]["val"]
        TotalMonthlyLeavesCount = 0
        TotalTimesLate = 0  ##number of times the emp was late in a month

        avoidkeys = set(['monthEnd', 'monthStart'])
        ## this is done to avoid the monthEnd and monthStart being considered as employee names.
        ## Since these are the keys of the main dataset when extracted from emp timesheet excel
        empnames = set(excelEmployeeParsedData.keys()).difference(avoidkeys)

        self.__printFalg and print("Difference ", set(excelEmployeeParsedData.keys()).difference(avoidkeys))
        self.__printFalg and print("shopConfiguration:", configuredDates)
        self.__printFalg and print("employeeDetails", employeeDetails)
        dateTimeFormat = '%Y-%m-%d %H:%M:%S'
        dateFormat = '%Y-%m-%d'
        # loop through the extracted data of employees
        for empname in empnames:

            # Create new Employee dict
            empobj = {
                "name": "",
                # "monthStart": "",
                # "monthEnd": "",
                "totalMonthlySalary": 0,
                "dailySalary": 0,
                "CalculatedSalary": 0,
                "totalSundaysAbsent": 0,
                "totalHolidaytaken": 0,
                "totalHolidaysAllowed": 0,
                "totalHolidaysWorked": 0,
                "totalWorkdays": 0,
                "totalTimesLate": 0,
                "totalAdditions": 0,
                "totaldeductions": 0,
                "workhours": [
                ]
            }

            # Variables to hold the salary calculations
            totaldeductions = ""
            totalAdditions = ""

            ##empty the list for the employee
            del empobj["workhours"][:]

            empobj["name"] = empname
            empobj["totalHolidaysAllowed"] = TotalAllowedLeavesPerMonth
            # empobj["monthStart"] = excelEmployeeParsedData["monthStart"]
            # empobj["monthEnd"] = excelEmployeeParsedData["monthEnd"]

            empobj["totalWorkdays"] = (
                        datetime.datetime.strptime(excelEmployeeParsedData["monthEnd"], "%d/%m/%Y")
                        - datetime.datetime.strptime(excelEmployeeParsedData["monthStart"],
                                                     "%d/%m/%Y")).days

            # Check if employee exists in the database in order to check the shift timings and salary
            if empname in employeeDetails:
                self.__printFalg and print("\n\n\nEmp Found in DB", empobj["name"], employeeDetails[empname],
                      employeeDetails[empname]['shift'])
                # Update the employees current monthly salary
                empobj["totalMonthlySalary"] = employeeDetails[empname]['salary']
                empobj["dailySalary"] = round(empobj["totalMonthlySalary"] / empobj["totalWorkdays"] , 2)
                # if employee details found in database then loop through all the days the employee was on job
                # Set the start and end time for the shift for the day based on if the day is defined as special day in config excel or not
                for empInoutTimeIndex in excelEmployeeParsedData[empname]['puchtime']:
                    dailyhours = {
                        "day": "",
                        "dayOfWeek": "",
                        "shiftStart": "",
                        "shiftEnd": "",
                        "shiftType": "",
                        "shiftComment": "",
                        "EmpTimeIn": "",
                        "EmpTimeOut": "",
                        "EnforcedEmpTimeIn": "",
                        "totalHoursLate": 0,
                        "PercentDeductions": 0,
                        "overTimeHours": 0,
                        "PercentSalaryAdditions": 0,
                        "halfDay": 0,
                        "Absent": 0,
                        "EnforcedEmpTimeOut": "",
                        "HolidayCredit": 0,
                        "DailyDeductions" : 0,
                        "DailyAdditions" : 0
                    }

                    # variables
                    ## used to set the emp start time to shop start time in case he comes early
                    empActualStart = ""

                    ## total number of hours deducted for employee based on various criteria
                    totalPercentSalaryDeducted = 0

                    # total amount of salary in percentage (of a day) the employee gained due to extra work criteria
                    totalPercentSalaryAcrued = 0

                    # total number of hours the employee is late ( grace of 5 minutes)
                    totalDailyHoursLate = 0
                    totalDailyOverTime = 0

                    shiftName = employeeDetails[empname]['shift']
                    nondefaultSearchStr = empInoutTimeIndex['day'].strftime(dateFormat) + ":" + shiftName
                    shopstartTime = None
                    shopendTime = None

                    # Check if the day in consideration is a special day when shop opened late of a holiday based on shop configuration excel tab
                    # check non default shift time if defined in the shop config
                    # Search for the day and shift name combination in non default configuration defined from excel
                    if nondefaultSearchStr in configuredDates.keys():
                        shopstartTime = configuredDates[nondefaultSearchStr]['intime']
                        shopendTime = configuredDates[nondefaultSearchStr]['outtime']
                        self.__printFalg and print("Non Default Shift:", nondefaultSearchStr, "shopstartTime", shopstartTime, shopendTime,
                              "EmpTime:", empInoutTimeIndex['startTime'], empInoutTimeIndex['endTime'])
                        dailyhours["shiftComment"] = configuredDates[nondefaultSearchStr]["comments"]
                    else:
                        # Its not a special day, based on the shift assigned to employee in config excel get the default start and end time from cnfgjson
                        if employeeDetails[empname]['shift'] in cnfgjson.keys():
                            shopstartTime = parse(
                                empInoutTimeIndex['day'].strftime(dateFormat) + " " + cnfgjson[shiftName]["OpenTime"])
                            shopendTime = parse(
                                empInoutTimeIndex['day'].strftime(dateFormat) + " " + cnfgjson[shiftName]["CloseTime"])
                            self.__printFalg and print("ShiftTime:", shopstartTime, shopendTime,
                                  "EmpTime:", empInoutTimeIndex['startTime'], empInoutTimeIndex['endTime'])
                        else:
                            ## in tis case we have not defined the default shift schedule in cnfgjson file
                            self.__printFalg and print(empname, employeeDetails[empname]['shift'], "not found in cnfgjson")
                            shopstartTime = "NA"
                            shopendTime = "NA"
                            raise ValueError(
                                'could not find {} in {}'.format(employeeDetails[empname]['shift'], cnfgjson.keys()))

                    # If employee forgets to set his outtime, then mark the day as half day and use empOuttime in calculations
                    if (empInoutTimeIndex['endTime'] is None or \
                        empInoutTimeIndex['endTime'] == "" or \
                        empInoutTimeIndex['endTime'] == parse(
                                empInoutTimeIndex['day'].strftime(dateFormat) + " 00:00:00")) \
                            and empInoutTimeIndex['startTime'] > shopstartTime:
                        dailyhours["halfDay"] = 1

                        empOuttime = shopendTime - datetime.timedelta(
                            seconds=int((shopendTime - shopstartTime).seconds / 2))
                        dailyhours["EnforcedEmpTimeOut"] = empOuttime
                        if empInoutTimeIndex['startTime'] > dailyhours["EnforcedEmpTimeOut"]:
                            empOuttime = shopendTime
                    else:
                        empOuttime = empInoutTimeIndex['endTime']

                    dailyhours["shiftStart"] = shopstartTime
                    dailyhours["shiftEnd"] = shopendTime
                    dailyhours["shiftType"] = employeeDetails[empname]['shift']
                    dailyhours["EmpTimeIn"] = empInoutTimeIndex['startTime']
                    dailyhours["EmpTimeOut"] = empInoutTimeIndex['endTime']
                    dailyhours["day"] = empInoutTimeIndex['day']
                    dailyhours["dayOfWeek"] = empInoutTimeIndex["dayOfWeek"]
                    dailyhours["shiftWorkHours"] = (shopendTime - shopstartTime).seconds // 3600

                    # late calculations
                    shopEmpTimeDelta = empInoutTimeIndex[ 'startTime'] - shopstartTime  ## emp start time minus shop start time
                    empStartEndTimDelta = empOuttime - empInoutTimeIndex[
                        'startTime']  ## emp end time minus emp start tim, to check if employee was absent
                    # if the employee start and end time difference is less than 60 minutes
                    if empStartEndTimDelta.days <= 0 and empStartEndTimDelta.seconds / 60 < 60:
                        self.__printFalg and print("Employee Absent", empInoutTimeIndex['dayOfWeek'], "days value:",
                              empStartEndTimDelta.days, "minutes:", empStartEndTimDelta.seconds / 60)
                        empobj["totalHolidaytaken"] = empobj["totalHolidaytaken"] + 1
                        dailyhours["Absent"] = 1
                        if empInoutTimeIndex['dayOfWeek'] == 'Sun':
                            empobj["totalSundaysAbsent"] = empobj["totalSundaysAbsent"] + 1
                    else:
                        # if time difference between shop open time and employee IN time is negative, meaning employee entered before his shift time
                        if empInoutTimeIndex['startTime'] < shopstartTime:
                            empActualStart = shopstartTime
                            dailyhours["EnforcedEmpTimeIn"] = shopstartTime
                            self.__printFalg and print("Too early....", "empInTime:", empInoutTimeIndex['startTime'], "shopstart:",
                                  shopstartTime, "empActualStart:", empActualStart)
                        # If employee came after store opening hours then calculate hours late
                        elif empInoutTimeIndex['startTime'] >= shopstartTime:
                            totalDailyHoursLate = (shopEmpTimeDelta.seconds / 60) // 60
                            if round((shopEmpTimeDelta.seconds / 60) % 60, 0) > inTimeGracePeriod:
                                totalDailyHoursLate = totalDailyHoursLate + 1

                            dailyhours["totalHoursLate"] = totalDailyHoursLate
                            if dailyhours["totalHoursLate"] >= 1:
                                empobj["totalTimesLate"] = empobj["totalTimesLate"] + 1
                                self.__printFalg and print("Times Late ............:", empobj["totalTimesLate"])

                            dailyhours["PercentDeductions"] = totalDailyHoursLate * hourlyDeductionPercent
                            dailyhours["DailyDeductions"] = round((empobj["dailySalary"] * dailyhours["PercentDeductions"])/100,2)
                            # In case if employee comes towards the shift and swipes in the card
                            if dailyhours["totalHoursLate"] >= 10:
                                dailyhours["halfDay"] = 0
                                dailyhours["Absent"] = 1
                                empobj["totalHolidaytaken"] = empobj["totalHolidaytaken"] + 1

                            self.__printFalg and print("Late Calculations: ", "hours Late:", (shopEmpTimeDelta.seconds / 60) // 60,
                                  "Minutes Late:", round((shopEmpTimeDelta.seconds / 60) % 60, 2),
                                  "totalDailyHoursLate", totalDailyHoursLate, "totalPercentSalaryDeducted",
                                  totalPercentSalaryDeducted)
                        else:
                            self.__printFalg and print("No Criteria Met:", "ShopOpen:", shopstartTime, "EmpTIme:",
                                  empInoutTimeIndex['startTime'])

                    # This check is for case when it is a holiday, special to check if staff is eligible for additional salary
                    holidaySearchStr = empInoutTimeIndex['day'].strftime(dateFormat) + ":" + "holiday"
                    if holidaySearchStr in configuredDates.keys():
                        holidayFlag = True
                        dailyhours["shiftType"] = configuredDates[holidaySearchStr]['entryType']
                        if dailyhours["Absent"] != 1:
                            empobj["totalHolidaysWorked"] = empobj["totalHolidaysWorked"] + 1
                            dailyhours["HolidayCredit"] = dailyhours["HolidayCredit"] + 1

                            self.__printFalg and print("Holiday Credit Done:", "totalHolidaysWorked", empobj["totalHolidaysWorked"],
                              "dailyhours[HolidayCredit]", dailyhours["HolidayCredit"])

                    empobj["workhours"].append(dailyhours)

                    ## if employee is doing extra time based on his shift
                    if empOuttime > shopendTime:
                        overtime = empOuttime - shopendTime
                        totalDailyOverTime = (overtime.seconds / 60) // 60  ##hours
                        if round((overtime.seconds / 60) % 60, 0) > overTimeLimit :  ## if remainder is over 30 minutes
                            totalDailyOverTime = totalDailyOverTime + 1
                        dailyhours["overTimeHours"] = totalDailyOverTime
                        dailyhours["PercentSalaryAdditions"] = dailyhours["overTimeHours"] * (hourlyAdditionPercent)/100
                        dailyhours["DailyAdditions"] = round((empobj["dailySalary"] * dailyhours["PercentSalaryAdditions"]), 2)
                        self.__printFalg and print("OverTime added:", overtime,  dailyhours["overTimeHours"], dailyhours["PercentSalaryAdditions"])

                empobj = self.applyComputeSalaryLogic(empobj)

                compiledSalaries[empobj["name"]] = empobj

            else:
                print("Emp Details Not Found in Database, Update employee Details in DB", empname)
                # compiledSalaries['err'].append('applyBuinessRules: could not find {} in database'.format(empname))
                empobj["name"] = "Could Not Find " + empname + " in Database, Insert Employee Details in DB"
                compiledSalaries[empname] = empobj

                # raise ValueError(
                #     'could not find {} in database'.format(empname))

        return compiledSalaries
